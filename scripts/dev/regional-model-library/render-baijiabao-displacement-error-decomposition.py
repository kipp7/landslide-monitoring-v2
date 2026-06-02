from __future__ import annotations

import json
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_OUT_DIR = Path(
    r"E:\学校\02 项目\04 各种比赛\03 计算机大赛\02_山体滑坡\测试与证明材料\04_AI模型测试证明\quantified-model-materials\error-decomposition"
)

plt.rcParams.update(
    {
        "font.family": "DejaVu Sans",
        "axes.titleweight": "bold",
        "axes.labelsize": 10,
        "axes.titlesize": 13,
        "figure.dpi": 140,
        "savefig.dpi": 220,
        "axes.spines.top": False,
        "axes.spines.right": False,
    }
)


def pct(value: float | int | None) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{float(value) * 100:.2f}%"


def fmt(value: float | int | None, digits: int = 3) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{float(value):.{digits}f}"


def load_tables(out_dir: Path) -> dict[str, pd.DataFrame]:
    names = [
        "per-sample-predictions",
        "skipped-samples",
        "byPoint",
        "byMonth",
        "bySeason",
        "byPointMonth",
        "byRainfall24h",
        "byRainfall72h",
        "byReservoirTrend72h",
        "byDisplacementTrend72h",
        "byDisplacementDelta72h",
        "byLabelMagnitude",
    ]
    return {name: pd.read_csv(out_dir / f"{name}.csv") for name in names}


def save_fig(fig: plt.Figure, charts_dir: Path, name: str) -> None:
    charts_dir.mkdir(parents=True, exist_ok=True)
    fig.savefig(charts_dir / f"{name}.png", bbox_inches="tight", facecolor="white")
    fig.savefig(charts_dir / f"{name}.svg", bbox_inches="tight", facecolor="white")
    plt.close(fig)


def add_value_labels(ax: plt.Axes, bars, digits: int = 2) -> None:
    for bar in bars:
        height = bar.get_height()
        if not np.isfinite(height):
            continue
        ax.text(
            bar.get_x() + bar.get_width() / 2,
            height,
            f"{height:.{digits}f}",
            ha="center",
            va="bottom",
            fontsize=8,
            color="#334155",
        )


def render_dashboard(summary: dict, charts_dir: Path) -> None:
    metrics = summary["overallMetrics"]
    fig = plt.figure(figsize=(11, 6.4))
    gs = fig.add_gridspec(2, 4, height_ratios=[1.05, 1.3], hspace=0.35, wspace=0.25)
    fig.suptitle("Baijiabao 24h Displacement Forecast - Error Decomposition Overview", x=0.03, ha="left", fontsize=16)

    cards = [
        ("MAE", f"{metrics['mae']:.3f} mm", "#0f766e"),
        ("RMSE", f"{metrics['rmse']:.3f} mm", "#0369a1"),
        ("R2", f"{metrics['r2']:.4f}", "#7c2d12"),
        ("Within 1mm", pct(metrics["within1mm"]), "#166534"),
    ]
    for idx, (label, value, color) in enumerate(cards):
        ax = fig.add_subplot(gs[0, idx])
        ax.set_facecolor("#f8fafc")
        ax.text(0.06, 0.68, label, fontsize=11, color="#475569", transform=ax.transAxes)
        ax.text(0.06, 0.30, value, fontsize=22, color=color, fontweight="bold", transform=ax.transAxes)
        ax.set_xticks([])
        ax.set_yticks([])
        for spine in ax.spines.values():
            spine.set_visible(False)

    ax = fig.add_subplot(gs[1, :])
    names = ["Direction", "Within 1mm", "Threshold state", "Threshold precision"]
    values = [
        metrics["directionAccuracy"],
        metrics["within1mm"],
        metrics["thresholdAgreement"],
        metrics["thresholdPrecision"],
    ]
    colors = ["#475569", "#0f766e", "#0369a1", "#b45309"]
    bars = ax.bar(names, values, color=colors, width=0.55)
    ax.set_ylim(0, 1.05)
    ax.set_ylabel("Rate")
    ax.set_title("Operational Agreement Metrics")
    ax.grid(axis="y", color="#e2e8f0", linewidth=0.8)
    for bar, value in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2, value + 0.025, pct(value), ha="center", fontsize=9)
    save_fig(fig, charts_dir, "00-error-decomposition-dashboard")


def render_point_chart(by_point: pd.DataFrame, charts_dir: Path) -> None:
    data = by_point.sort_values("pointId").copy()
    fig, ax = plt.subplots(figsize=(9, 5.2))
    x = np.arange(len(data))
    width = 0.36
    bars1 = ax.bar(x - width / 2, data["mae"], width, label="MAE", color="#0f766e")
    bars2 = ax.bar(x + width / 2, data["rmse"], width, label="RMSE", color="#0369a1")
    add_value_labels(ax, bars1)
    add_value_labels(ax, bars2)
    ax.set_xticks(x)
    ax.set_xticklabels(data["pointId"])
    ax.set_ylabel("Error (mm)")
    ax.set_title("MAE / RMSE by Monitoring Point")
    ax.legend(frameon=False)
    ax.grid(axis="y", color="#e2e8f0")
    save_fig(fig, charts_dir, "01-mae-rmse-by-point")


def render_month_chart(by_month: pd.DataFrame, charts_dir: Path) -> None:
    data = by_month.sort_values("month").copy()
    fig, ax1 = plt.subplots(figsize=(10, 5.4))
    x = np.arange(len(data))
    ax1.plot(x, data["mae"], marker="o", color="#0f766e", linewidth=2.3, label="MAE")
    ax1.plot(x, data["rmse"], marker="s", color="#0369a1", linewidth=2.3, label="RMSE")
    ax1.set_xticks(x)
    ax1.set_xticklabels(data["month"].astype(str).str.zfill(2))
    ax1.set_ylabel("Error (mm)")
    ax1.set_xlabel("Month")
    ax1.set_title("Monthly Error Profile")
    ax1.grid(axis="y", color="#e2e8f0")
    ax2 = ax1.twinx()
    ax2.bar(x, data["count"], alpha=0.18, color="#64748b", label="Count")
    ax2.set_ylabel("Sample count")
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, frameon=False, loc="upper left")
    save_fig(fig, charts_dir, "02-error-by-month")


def render_heatmap(by_point_month: pd.DataFrame, charts_dir: Path) -> None:
    pivot = by_point_month.pivot_table(index="pointId", columns="month", values="mae", aggfunc="mean")
    pivot = pivot.reindex(sorted(pivot.index), axis=0).reindex(sorted(pivot.columns), axis=1)
    fig, ax = plt.subplots(figsize=(11, 4.8))
    cmap = LinearSegmentedColormap.from_list("error_cmap", ["#ecfdf5", "#fde68a", "#f97316", "#7f1d1d"])
    im = ax.imshow(pivot.values, aspect="auto", cmap=cmap)
    ax.set_xticks(np.arange(len(pivot.columns)))
    ax.set_xticklabels([str(col).zfill(2) for col in pivot.columns])
    ax.set_yticks(np.arange(len(pivot.index)))
    ax.set_yticklabels(pivot.index)
    ax.set_xlabel("Month")
    ax.set_ylabel("Point")
    ax.set_title("Point x Month MAE Heatmap")
    for i in range(pivot.shape[0]):
        for j in range(pivot.shape[1]):
            value = pivot.values[i, j]
            if np.isfinite(value):
                ax.text(j, i, f"{value:.2f}", ha="center", va="center", fontsize=8, color="#111827")
    fig.colorbar(im, ax=ax, fraction=0.025, pad=0.02, label="MAE (mm)")
    save_fig(fig, charts_dir, "03-point-month-mae-heatmap")


def render_abs_error_distribution(samples: pd.DataFrame, charts_dir: Path) -> None:
    fig, ax = plt.subplots(figsize=(9.5, 5.2))
    ax.hist(samples["abs_error"], bins=36, color="#0f766e", alpha=0.82, edgecolor="white")
    p50 = samples["abs_error"].quantile(0.5)
    p90 = samples["abs_error"].quantile(0.9)
    ax.axvline(p50, color="#0369a1", linewidth=2, label=f"P50={p50:.3f}mm")
    ax.axvline(p90, color="#b45309", linewidth=2, label=f"P90={p90:.3f}mm")
    ax.axvline(1.0, color="#7f1d1d", linewidth=1.8, linestyle="--", label="1mm tolerance")
    ax.set_xlabel("Absolute error (mm)")
    ax.set_ylabel("Sample count")
    ax.set_title("Absolute Error Distribution")
    ax.grid(axis="y", color="#e2e8f0")
    ax.legend(frameon=False)
    save_fig(fig, charts_dir, "04-absolute-error-distribution")


def render_true_vs_pred(samples: pd.DataFrame, charts_dir: Path) -> None:
    fig, ax = plt.subplots(figsize=(7.2, 6.4))
    colors = np.where(samples["within_1mm"].astype(str).str.lower().eq("true"), "#0f766e", "#b45309")
    ax.scatter(samples["y_true"], samples["y_pred"], s=18, c=colors, alpha=0.62, edgecolors="none")
    low = float(min(samples["y_true"].min(), samples["y_pred"].min()))
    high = float(max(samples["y_true"].max(), samples["y_pred"].max()))
    ax.plot([low, high], [low, high], color="#334155", linewidth=1.6, linestyle="--", label="Ideal")
    ax.axhline(0, color="#cbd5e1", linewidth=1)
    ax.axvline(0, color="#cbd5e1", linewidth=1)
    ax.set_xlabel("Observed 24h displacement delta (mm)")
    ax.set_ylabel("Predicted 24h displacement delta (mm)")
    ax.set_title("Observed vs Predicted Displacement")
    ax.legend(frameon=False)
    ax.grid(color="#e2e8f0")
    save_fig(fig, charts_dir, "05-true-vs-predicted-scatter")


def render_residual_regime(samples: pd.DataFrame, charts_dir: Path) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(12, 5.2))
    axes[0].scatter(samples["rainfall_72h_mm"], samples["error"], s=16, alpha=0.55, color="#0369a1", edgecolors="none")
    axes[0].axhline(0, color="#334155", linewidth=1.3)
    axes[0].set_xlabel("72h rainfall (mm)")
    axes[0].set_ylabel("Prediction error (mm)")
    axes[0].set_title("Residual vs Rainfall")
    axes[0].grid(color="#e2e8f0")
    axes[1].scatter(samples["reservoir_delta_72h_m"], samples["error"], s=16, alpha=0.55, color="#b45309", edgecolors="none")
    axes[1].axhline(0, color="#334155", linewidth=1.3)
    axes[1].set_xlabel("72h reservoir level delta (m)")
    axes[1].set_ylabel("Prediction error (mm)")
    axes[1].set_title("Residual vs Reservoir Trend")
    axes[1].grid(color="#e2e8f0")
    save_fig(fig, charts_dir, "06-residual-vs-rainfall-reservoir")


def render_regime_bars(tables: dict[str, pd.DataFrame], charts_dir: Path) -> None:
    panels = [
        ("byRainfall72h", "rainfall_72h_bucket", "72h Rainfall Regime"),
        ("byReservoirTrend72h", "reservoir_trend_72h", "Reservoir Trend"),
        ("byDisplacementTrend72h", "displacement_trend_72h", "Displacement Trend"),
        ("byLabelMagnitude", "label_abs_bucket", "Label Magnitude"),
    ]
    fig, axes = plt.subplots(2, 2, figsize=(12.5, 8))
    for ax, (table_name, key, title) in zip(axes.flat, panels):
        data = tables[table_name].sort_values(key)
        bars = ax.bar(data[key].astype(str), data["mae"], color="#0f766e", alpha=0.88)
        add_value_labels(ax, bars, digits=2)
        ax.set_title(title)
        ax.set_ylabel("MAE (mm)")
        ax.tick_params(axis="x", rotation=25)
        ax.grid(axis="y", color="#e2e8f0")
    fig.suptitle("Error by Hydrologic and Displacement Regimes", fontsize=15, fontweight="bold", x=0.03, ha="left")
    fig.tight_layout()
    save_fig(fig, charts_dir, "07-error-by-regime-bars")


def render_charts(tables: dict[str, pd.DataFrame], summary: dict, charts_dir: Path) -> None:
    render_dashboard(summary, charts_dir)
    render_point_chart(tables["byPoint"], charts_dir)
    render_month_chart(tables["byMonth"], charts_dir)
    render_heatmap(tables["byPointMonth"], charts_dir)
    render_abs_error_distribution(tables["per-sample-predictions"], charts_dir)
    render_true_vs_pred(tables["per-sample-predictions"], charts_dir)
    render_residual_regime(tables["per-sample-predictions"], charts_dir)
    render_regime_bars(tables, charts_dir)


def write_workbook(tables: dict[str, pd.DataFrame], summary: dict, out_dir: Path) -> Path:
    workbook_path = out_dir / "Baijiabao位移预测误差分解数据包.xlsx"
    overall = summary["overallMetrics"]
    overview = pd.DataFrame(
        [
            ["model", summary["model"]["displayName"]],
            ["modelKey", summary["model"]["modelKey"]],
            ["evaluatedCount", summary["evaluatedCount"]],
            ["skippedCount", summary["skippedCount"]],
            ["MAE_mm", overall["mae"]],
            ["RMSE_mm", overall["rmse"]],
            ["R2", overall["r2"]],
            ["DirectionAccuracy", overall["directionAccuracy"]],
            ["Within1mm", overall["within1mm"]],
            ["ThresholdStateAgreement", overall["thresholdAgreement"]],
            ["P50AbsError_mm", overall["p50AbsError"]],
            ["P90AbsError_mm", overall["p90AbsError"]],
        ],
        columns=["Metric", "Value"],
    )
    dictionary = pd.DataFrame(
        [
            ["per-sample-predictions", "逐样本真实值、预测值、残差和分组标签主表"],
            ["byPoint", "按 GNSS 监测点位聚合的误差分解"],
            ["byMonth", "按月份聚合的误差分解"],
            ["byPointMonth", "点位 x 月份误差热力图数据"],
            ["byRainfall24h / byRainfall72h", "按降雨窗口分箱的误差分解"],
            ["byReservoirTrend72h", "按 72h 水库水位涨落状态分解"],
            ["byDisplacementTrend72h", "按 72h 位移趋势分解"],
            ["byLabelMagnitude", "按真实 24h 位移增量幅值分解"],
            ["skipped-samples", "因运行时必需特征缺失而跳过的验证样本"],
        ],
        columns=["Sheet/Table", "Meaning"],
    )
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        dictionary.to_excel(writer, sheet_name="DataDictionary", index=False)
        sheet_map = {
            "per-sample-predictions": "PerSample",
            "skipped-samples": "Skipped",
            "byPoint": "ByPoint",
            "byMonth": "ByMonth",
            "bySeason": "BySeason",
            "byPointMonth": "ByPointMonth",
            "byRainfall24h": "ByRainfall24h",
            "byRainfall72h": "ByRainfall72h",
            "byReservoirTrend72h": "ByReservoir",
            "byDisplacementTrend72h": "ByDispTrend",
            "byDisplacementDelta72h": "ByDispDelta",
            "byLabelMagnitude": "ByLabelMag",
        }
        for name, sheet_name in sheet_map.items():
            tables[name].to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(workbook_path)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for column_cells in ws.columns:
            letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells[:200])
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)
        if ws.max_row > 2:
            headers = [cell.value for cell in ws[1]]
            for metric in ["mae", "rmse", "abs_error", "p90AbsError"]:
                if metric in headers:
                    col = get_column_letter(headers.index(metric) + 1)
                    ws.conditional_formatting.add(
                        f"{col}2:{col}{ws.max_row}",
                        ColorScaleRule(
                            start_type="min",
                            start_color="ECFDF5",
                            mid_type="percentile",
                            mid_value=50,
                            mid_color="FDE68A",
                            end_type="max",
                            end_color="F97316",
                        ),
                    )
    wb.save(workbook_path)
    return workbook_path


def update_manifest(out_dir: Path, workbook_path: Path, charts_dir: Path) -> None:
    manifest_path = out_dir / "manifest.csv"
    manifest = pd.read_csv(manifest_path)
    extras = [
        {
            "file": workbook_path.name,
            "description": "Excel workbook containing all error decomposition tables",
            "rows": 1,
        }
    ]
    for file in sorted(charts_dir.glob("*.png")):
        extras.append({"file": f"charts/{file.name}", "description": "Error decomposition chart PNG", "rows": 1})
    for file in sorted(charts_dir.glob("*.svg")):
        extras.append({"file": f"charts/{file.name}", "description": "Error decomposition chart SVG", "rows": 1})
    manifest = pd.concat([manifest, pd.DataFrame(extras)], ignore_index=True)
    manifest.to_csv(manifest_path, index=False, encoding="utf-8-sig")
    (out_dir / "manifest.json").write_text(
        json.dumps(manifest.to_dict(orient="records"), ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    out_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUT_DIR
    charts_dir = out_dir / "charts"
    tables = load_tables(out_dir)
    summary = json.loads((out_dir / "error-decomposition-summary.json").read_text(encoding="utf-8"))
    render_charts(tables, summary, charts_dir)
    workbook_path = write_workbook(tables, summary, out_dir)
    update_manifest(out_dir, workbook_path, charts_dir)
    print(f"Rendered workbook: {workbook_path}")
    print(f"Rendered charts: {charts_dir}")


if __name__ == "__main__":
    main()
